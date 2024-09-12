import streamlit as st
import pandas as pd
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import time
import imaplib
import email
import re
from openpyxl import load_workbook

# Function to send emails
def send_bulk_emails(sender_email, sender_password, subject, body, recipients):
    smtp_server = 'smtp.gmail.com'
    smtp_port = 587

    server = smtplib.SMTP(smtp_server, smtp_port)
    server.starttls()
    server.login(sender_email, sender_password)

    progress_bar = st.progress(0)
    progress_text = st.empty()
    total_emails = len(recipients)
    successful_sends = 0
    failed_sends = 0

    for idx, recipient in enumerate(recipients):
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient
        msg['Subject'] = subject
        msg.attach(MIMEText(body, 'plain'))

        try:
            server.sendmail(sender_email, recipient, msg.as_string())
            successful_sends += 1
        except Exception as e:
            failed_sends += 1

        progress = (idx + 1) / total_emails
        progress_bar.progress(min(progress, 1.0))
        progress_text.text(f"Sent {idx + 1} out of {total_emails} emails")

    server.quit()

    st.success(f"Emails sent successfully to {successful_sends} recipients.")
    if failed_sends > 0:
        st.error(f"Failed to send emails to {failed_sends} recipients.")

# Function to get email addresses from file
def get_email_addresses_from_file(uploaded_file, column_name):
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
    email_addresses = df[column_name].dropna().str.strip().tolist()  # Remove blank spaces
    return email_addresses

# Function to get email subjects and bodies from file
def get_subjects_and_bodies_from_file(uploaded_file, subject_column, body_column):
    df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
    subjects = df[subject_column].dropna().tolist()
    bodies = df[body_column].dropna().tolist()
    return subjects, bodies

# Function to fetch undelivered emails using IMAP
def fetch_undelivered_emails(email_address, app_password, recipients):
    imap_server = 'imap.gmail.com'
    imap_port = 993
    mailbox = 'INBOX'

    undelivered_emails = []
    try:
        mail = imaplib.IMAP4_SSL(imap_server, imap_port)
        mail.login(email_address, app_password)
        mail.select(mailbox)

        # Search for "Mail Delivery Subsystem" emails
        status, response = mail.search(None, 'FROM', '"Mail Delivery Subsystem"')
        email_ids = response[0].split()

        for e_id in email_ids:
            status, msg_data = mail.fetch(e_id, '(RFC822)')
            for response_part in msg_data:
                if isinstance(response_part, tuple):
                    msg = email.message_from_bytes(response_part[1])
                    if msg.is_multipart():
                        for part in msg.walk():
                            if part.get_content_type() == 'text/plain':
                                email_body = part.get_payload(decode=True).decode('utf-8')
                                match = re.search(r'Your message wasn\'t delivered to ([\w.-]+@[\w.-]+)', email_body)
                                if match and match.group(1) in recipients:
                                    undelivered_emails.append(match.group(1))
                    else:
                        email_body = msg.get_payload(decode=True).decode('utf-8')
                        match = re.search(r'Your message wasn\'t delivered to ([\w.-]+@[\w.-]+)', email_body)
                        if match and match.group(1) in recipients:
                            undelivered_emails.append(match.group(1))

        mail.logout()
    except Exception as e:
        st.error(f"Error fetching undelivered emails: {str(e)}")

    return undelivered_emails

# Function to update email addresses file based on undelivered emails
def update_email_addresses_file(original_file, undelivered_emails):
    is_csv = original_file.name.endswith('.csv')
    df = pd.read_csv(original_file) if is_csv else pd.read_excel(original_file)
    wb = load_workbook(original_file)
    ws = wb.active

    # Remove undelivered emails from the DataFrame
    df = df[~df['Email Address'].isin(undelivered_emails)]

    # Add a new column for undelivered emails
    spam_df = pd.DataFrame({'Spam Email Address': undelivered_emails})
    df = pd.concat([df, spam_df], axis=1)

    # Save to CSV or Excel file based on original file format
    updated_filename = 'updated_email_addresses.csv' if is_csv else 'updated_email_addresses.xlsx'
    if is_csv:
        df.to_csv(updated_filename, index=False)
    else:
        # Update the workbook with the new DataFrame content
        for idx, row in df.iterrows():
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=idx + 2, column=col_idx, value=value)
        # Add the new column header and values
        ws.cell(row=1, column=len(df.columns), value='Spam Email Address')
        for row_idx, email in enumerate(undelivered_emails, 2):
            ws.cell(row=row_idx, column=len(df.columns), value=email)
        wb.save(updated_filename)

    return updated_filename

# Streamlit app
st.title('Bulk Email Sender & Undelivered Email Checker')

# Upload email addresses file
st.subheader('Upload Email Addresses File')
uploaded_emails_file = st.file_uploader("Upload a CSV or Excel file with email addresses", type=['csv', 'xlsx'], key="emails")

# Upload email contents file
st.subheader('Upload Email Contents File')
uploaded_contents_file = st.file_uploader("Upload a CSV or Excel file with email subjects and bodies", type=['csv', 'xlsx'], key="contents")

# Track whether undelivered emails have been processed
undelivered_processed = False
updated_filename = None

if uploaded_emails_file and uploaded_contents_file:
    # Display the content of the uploaded email addresses file
    st.write("Email Addresses File Content:")
    email_df = pd.read_csv(uploaded_emails_file) if uploaded_emails_file.name.endswith('.csv') else pd.read_excel(uploaded_emails_file)
    st.write(email_df)

    # Display the content of the uploaded email contents file
    st.write("Email Contents File Content:")
    content_df = pd.read_csv(uploaded_contents_file) if uploaded_contents_file.name.endswith('.csv') else pd.read_excel(uploaded_contents_file)
    st.write(content_df)

    # Input email details
    sender_email = 'spamchecker678@gmail.com'
    sender_password = 'hxzi umco hjug afsk'
    email_column = 'Email Address'
    subject_column = 'Subject'
    body_column = 'Body'

    if st.button('Send Emails'):
        if not sender_email or not sender_password or not email_column or not subject_column or not body_column:
            st.error("Please fill in all the fields.")
        else:
            recipients = get_email_addresses_from_file(uploaded_emails_file, email_column)
            subjects, bodies = get_subjects_and_bodies_from_file(uploaded_contents_file, subject_column, body_column)
            send_bulk_emails(sender_email, sender_password, subjects[0], bodies[0], recipients)

    if st.button("Start Monitoring Undelivered Emails"):
        st.write("Monitoring started. Please do not close this window for the next 1 minute.")
        recipients = get_email_addresses_from_file(uploaded_emails_file, email_column)

        undelivered_emails = []
        progress_bar = st.progress(0)
        progress_text = st.empty()
        start_time = time.time()
        monitoring_duration = 60  # 1 minute

        while time.time() - start_time <= monitoring_duration:
            undelivered_emails.extend(fetch_undelivered_emails(sender_email, sender_password, recipients))
            undelivered_emails = list(set(undelivered_emails))  # Remove duplicates
            time.sleep(10)
            progress = (time.time() - start_time) / monitoring_duration
            progress_bar.progress(min(progress, 1.0))
            progress_text.text(f"Monitoring progress: {int(progress * 100)}%")

        if undelivered_emails:
            updated_filename = update_email_addresses_file(uploaded_emails_file, undelivered_emails)
            st.success(f"Undelivered emails found and updated email addresses saved to '{updated_filename}'.")
            undelivered_processed = True
        else:
            st.warning("No undelivered emails found.")

# Provide download button for updated email addresses file
if undelivered_processed and updated_filename:
    file_format = 'csv' if updated_filename.endswith('.csv') else 'xlsx'
    st.download_button('Download Updated Email Addresses', data=open(updated_filename, 'rb').read(), file_name=updated_filename, mime=f'application/{file_format}')